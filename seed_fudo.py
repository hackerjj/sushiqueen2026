#!/usr/bin/env python3
"""Seed MongoDB Atlas with Fudo backup data"""
import openpyxl, xlrd, datetime
from pymongo import MongoClient

MONGO_URI = "mongodb+srv://jairalonsogarcias_db_user:rTS47frTdRtG2NUh@sushiqueen.34hq2xo.mongodb.net/sushi_queen?retryWrites=true&w=majority&appName=sushiqueen"
client = MongoClient(MONGO_URI)
db = client['sushi_queen']

def safe_str(v):
    if v is None: return ''
    if isinstance(v, float): return str(int(v)) if v == int(v) else str(v)
    return str(v).strip()

def safe_date(v):
    if isinstance(v, datetime.datetime): return v
    if isinstance(v, str):
        for fmt in ['%d/%m/%Y', '%Y-%m-%d']:
            try: return datetime.datetime.strptime(v, fmt)
            except: pass
    return None

# ─── 1. CLIENTES (902 rows) ──────────────────────────────────
print("=== Seeding Customers ===")
wb = openpyxl.load_workbook('DatosFudo/clientes-63956-20260305-1-i5hlt5.xlsx')
sh = wb.active
customers = []
for i, row in enumerate(sh.iter_rows(values_only=True)):
    if i == 0: continue  # header
    if not row[1]: continue
    customers.append({
        'fudo_id': row[0],
        'name': safe_str(row[1]),
        'email': safe_str(row[2]),
        'phone': safe_str(row[3]),
        'address': {
            'street': safe_str(row[4]),
            'number': safe_str(row[5]),
            'floor': safe_str(row[6]),
            'city': safe_str(row[7]),
        },
        'source': safe_str(row[8]) or 'fudo',
        'total_orders': int(row[9] or 0),
        'last_order': safe_str(row[10]),
        'balance': float(row[11] or 0),
        'payment_method': safe_str(row[12]),
        'discount': float(row[13] or 0),
        'notes': safe_str(row[14]),
        'active': safe_str(row[18]) == 'Sí',
    })
if customers:
    db.customers.delete_many({})
    db.customers.insert_many(customers)
    print(f"  Inserted {len(customers)} customers")

# ─── 2. PROVEEDORES (63 rows) ────────────────────────────────
print("=== Seeding Suppliers ===")
wb = openpyxl.load_workbook('DatosFudo/proveedores-63956-20260305-1-wnc3uh.xlsx')
sh = wb.active
suppliers = []
for i, row in enumerate(sh.iter_rows(values_only=True)):
    if i == 0: continue
    if not row[1]: continue
    suppliers.append({
        'fudo_id': row[0],
        'name': safe_str(row[1]),
        'email': safe_str(row[2]),
        'phone': safe_str(row[3]),
        'address': safe_str(row[4]),
        'tax_id': safe_str(row[5]),
        'balance': float(row[6] or 0),
        'notes': safe_str(row[7]),
        'active': safe_str(row[8]) == 'Sí',
    })
if suppliers:
    db.suppliers.delete_many({})
    db.suppliers.insert_many(suppliers)
    print(f"  Inserted {len(suppliers)} suppliers")

# ─── 3. INGREDIENTES (179 rows) ──────────────────────────────
print("=== Seeding Ingredients ===")
wb = xlrd.open_workbook('DatosFudo/ingredientes.xls')
sh = wb.sheet_by_index(0)
ingredients = []
for i in range(1, sh.nrows):
    row = [sh.cell_value(i, j) for j in range(sh.ncols)]
    if not row[2]: continue
    ingredients.append({
        'fudo_id': int(row[0]) if row[0] else None,
        'category': safe_str(row[1]),
        'name': safe_str(row[2]),
        'cost': float(row[3] or 0),
        'supplier': safe_str(row[4]),
        'unit': safe_str(row[5]),
        'stock_control': safe_str(row[6]) == 'Si',
        'stock': float(row[7] or 0),
        'waste': safe_str(row[8]),
        'min_stock': 5,
    })
if ingredients:
    db.ingredients.delete_many({})
    db.ingredients.insert_many(ingredients)
    print(f"  Inserted {len(ingredients)} ingredients")

# ─── 4. GASTOS (all files) ───────────────────────────────────
print("=== Seeding Expenses ===")
import glob
expenses = []
for f in glob.glob('DatosFudo/gastos-*.xlsx'):
    wb = openpyxl.load_workbook(f)
    sh = wb.active
    header_row = None
    for i, row in enumerate(sh.iter_rows(values_only=True)):
        if row and row[0] == 'Id':
            header_row = i
            continue
        if header_row is None: continue
        if not row[0]: continue
        expenses.append({
            'fudo_id': row[0],
            'date': row[1] if isinstance(row[1], datetime.datetime) else safe_date(safe_str(row[1])),
            'supplier': safe_str(row[2]),
            'category': safe_str(row[3]),
            'subcategory': safe_str(row[4]),
            'comment': safe_str(row[5]),
            'amount': float(row[6] or 0),
            'from_register': safe_str(row[7]) == 'Sí',
            'register': safe_str(row[8]),
            'payment_method': safe_str(row[9]),
            'cancelled': safe_str(row[14]) == 'Sí' if len(row) > 14 else False,
        })
if expenses:
    db.expenses.delete_many({})
    db.expenses.insert_many(expenses)
    print(f"  Inserted {len(expenses)} expenses")

# ─── 5. VENTAS (all Reporte-Ventas files) ─────────────────────
print("=== Seeding Sales Reports ===")
sales = []
for f in glob.glob('DatosFudo/Reporte-Ventas*.xlsx'):
    wb = openpyxl.load_workbook(f)
    sh = wb.active
    for i, row in enumerate(sh.iter_rows(values_only=True)):
        if i == 0: continue
        if not row[0]: continue
        sales.append({
            'date': row[0] if isinstance(row[0], datetime.datetime) else None,
            'day_of_week': row[1] if isinstance(row[1], datetime.datetime) else safe_str(row[1]),
            'hour': int(row[2]) if row[2] else None,
            'total': float(row[3] or 0),
            'count': int(row[4] or 0),
        })
if sales:
    db.sales_reports.delete_many({})
    db.sales_reports.insert_many(sales)
    print(f"  Inserted {len(sales)} sales records")

# ─── 6. MOVIMIENTOS DE CAJA (all files) ──────────────────────
print("=== Seeding Cash Movements ===")
movements = []
for f in glob.glob('DatosFudo/movimientos-de-caja*.xls'):
    wb = xlrd.open_workbook(f)
    sh = wb.sheet_by_index(0)
    for i in range(1, sh.nrows):
        row = [sh.cell_value(i, j) for j in range(sh.ncols)]
        if not row[0] or row[0] == 'ID' or row[0] == 'Id': continue
        try:
            movements.append({
                'fudo_id': int(row[0]) if isinstance(row[0], float) else row[0],
                'date': safe_str(row[1]),
                'type': safe_str(row[2]),
                'amount': float(row[3]) if len(row) > 3 and row[3] and isinstance(row[3], (int, float)) else 0,
                'register': safe_str(row[4]) if len(row) > 4 else '',
                'comment': safe_str(row[5]) if len(row) > 5 else '',
            })
        except: pass
if movements:
    db.cash_movements.delete_many({})
    db.cash_movements.insert_many(movements)
    print(f"  Inserted {len(movements)} cash movements")

# ─── 7. PROPINAS (all files) ──────────────────────────────────
print("=== Seeding Tips ===")
tips = []
for f in glob.glob('DatosFudo/propinas*.xls'):
    wb = xlrd.open_workbook(f)
    sh = wb.sheet_by_index(0)
    for i in range(1, sh.nrows):
        row = [sh.cell_value(i, j) for j in range(sh.ncols)]
        if not row[0] or safe_str(row[0]).lower() in ('id', 'Id', 'ID'): continue
        try:
            tips.append({
                'fudo_id': int(row[0]) if isinstance(row[0], float) else row[0],
                'date': safe_str(row[1]) if len(row) > 1 else '',
                'amount': float(row[2]) if len(row) > 2 and isinstance(row[2], (int, float)) else 0,
                'waiter': safe_str(row[3]) if len(row) > 3 else '',
            })
        except: pass
if tips:
    db.tips.delete_many({})
    db.tips.insert_many(tips)
    print(f"  Inserted {len(tips)} tips")

# ─── 8. VENTAS DETALLE (ventas*.xls) ─────────────────────────
print("=== Seeding Sales Detail ===")
sales_detail = []
for f in glob.glob('DatosFudo/ventas*.xls'):
    wb = xlrd.open_workbook(f)
    sh = wb.sheet_by_index(0)
    for i in range(1, sh.nrows):
        row = [sh.cell_value(i, j) for j in range(sh.ncols)]
        if not row[0] or safe_str(row[0]).lower() in ('id',): continue
        try:
            entry = {'fudo_id': int(row[0]) if isinstance(row[0], float) else row[0]}
            for j in range(1, sh.ncols):
                header = sh.cell_value(0, j) if sh.cell_value(0, j) else f'col_{j}'
                entry[header.lower().replace(' ', '_').replace('.', '')] = row[j]
            sales_detail.append(entry)
        except: pass
if sales_detail:
    db.sales_detail.delete_many({})
    db.sales_detail.insert_many(sales_detail)
    print(f"  Inserted {len(sales_detail)} sales detail records")

print("\n✅ All Fudo data seeded to MongoDB Atlas!")
client.close()
